"""Write SAH subject workbooks with the 7 core courseware sheets."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from schemas import (
    FORBIDDEN_SHEETS,
    QUESTIONS_SHEET_NAME,
    SAH_HEADERS,
    CHAPTER_MAP_HEADERS,
    TOPIC_MAP_HEADERS,
    LESSON_PLANS_HEADERS,
    CONCEPTS_HEADERS,
    HOMEWORK_HEADERS,
    RESOURCES_HEADERS,
    QuestionRow,
    ChapterMapRow,
    TopicMapRow,
    LessonPlanRow,
    ConceptRow,
    HomeworkRow,
    ResourceRow,
    SubjectDirectionPlan,
)
from validator import validate_question_rows

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_subject_workbook(
    output_path: Path,
    rows: list[QuestionRow],
    *,
    direction: SubjectDirectionPlan | None = None,
    chapter_map_rows: list[ChapterMapRow] | None = None,
    topic_map_rows: list[TopicMapRow] | None = None,
    lesson_plans_rows: list[LessonPlanRow] | None = None,
    concepts_rows: list[ConceptRow] | None = None,
    homework_rows: list[HomeworkRow] | None = None,
    resources_rows: list[ResourceRow] | None = None,
    enforce_maths_case_format: bool = False,
    validation_errors: list[str] | None = None,
) -> list[str]:
    """Writes the full 7-sheet SAH Courseware Workbook."""
    errors = validation_errors if validation_errors is not None else validate_question_rows(
        rows,
        direction=direction,
        enforce_maths_case_format=enforce_maths_case_format,
    )
    if errors:
        raise ValueError("Validation failed:\n" + "\n".join(f"  - {e}" for e in errors))

    # Resolve prefix (default to MTH9/SCI9 style)
    id_prefix = direction.id_prefix if direction else "MATH9"
    if rows and not id_prefix:
        # derive prefix from first question's ID, e.g. "MTH9-CH01-MCQ-001" -> "MATH9" or "MTH9"
        parts = rows[0].question_id.split("-")
        if len(parts) > 0:
            import re
            # match alphabetical and numeric suffix, e.g. SCI9 or MTH9
            m = re.match(r"^([a-zA-Z]+[0-9]+)", parts[0])
            if m:
                id_prefix = m.group(1)

    # 1. Resolve/Auto-generate sheet rows for complete workbook mapping
    resolved_chapters = chapter_map_rows
    resolved_topics = topic_map_rows
    resolved_lessons = lesson_plans_rows
    resolved_concepts = concepts_rows
    resolved_homework = homework_rows
    resolved_resources = resources_rows

    if resolved_chapters is None or resolved_topics is None:
        # Extract unique chapters and topics from questions to build maps
        chapters_dict: dict[int, str] = {}
        topics_by_ch: dict[int, list[str]] = {}
        for r in rows:
            chapters_dict[r.chapter_no] = r.chapter
            topics_by_ch.setdefault(r.chapter_no, [])
            if r.topic and r.topic not in topics_by_ch[r.chapter_no]:
                topics_by_ch[r.chapter_no].append(r.topic)

        # Generate Chapter_Map
        if resolved_chapters is None:
            resolved_chapters = []
            for ch_no, ch_title in sorted(chapters_dict.items()):
                ch_id = f"{id_prefix}_CH{ch_no:02d}"
                resolved_chapters.append(
                    ChapterMapRow(
                        chapter_id=ch_id,
                        chapter_no=ch_no,
                        chapter_title=ch_title,
                        default_priority=3,
                        status="active"
                    )
                )

        # Generate Topic_Map and other sub-sheets
        if resolved_topics is None:
            resolved_topics = []
            resolved_lessons = resolved_lessons or []
            resolved_concepts = resolved_concepts or []
            resolved_homework = resolved_homework or []
            resolved_resources = resolved_resources or []
            
            global_seq = 1
            for ch_no, ch_title in sorted(chapters_dict.items()):
                ch_id = f"{id_prefix}_CH{ch_no:02d}"
                ch_topics = topics_by_ch.get(ch_no, ["Introduction"])
                if not ch_topics:
                    ch_topics = ["Introduction"]

                for t_idx, t_name in enumerate(ch_topics, start=1):
                    t_id = f"{ch_id}_T{t_idx:02d}"
                    
                    # 1. Topic Map Row
                    resolved_topics.append(
                        TopicMapRow(
                            topic_id=t_id,
                            chapter_id=ch_id,
                            sequence_no=global_seq,
                            topic_title=t_name,
                            relative_weight=1.0,
                            relative_difficulty="Medium",
                            learning_outcomes=f"Understand key concepts of {t_name}",
                            status="active"
                        )
                    )

                    # 2. Lesson Plan Row
                    resolved_lessons.append(
                        LessonPlanRow(
                            lesson_plan_id=f"LP_CH{ch_no:02d}_T{t_idx:02d}",
                            chapter_id=ch_id,
                            topic_id=t_id,
                            objectives=f"Define core terms of {t_name};Solve fundamental problems",
                            phase_engage=f"Introduce {t_name} using practical life scenarios.",
                            phase_explore=f"Have students explore patterns and note observations.",
                            phase_explain=f"Formalize definitions and mathematical properties.",
                            phase_elaborate=f"Apply concepts to complex word problems.",
                            phase_evaluate=f"Conduct a 5-minute exit ticket checking understanding.",
                            required_resources="NCERT Textbook;Worksheet",
                            notes=""
                        )
                    )

                    # 3. Concept Row
                    resolved_concepts.append(
                        ConceptRow(
                            concept_id=f"CON_CH{ch_no:02d}_T{t_idx:02d}",
                            chapter_id=ch_id,
                            topic_id=t_id,
                            concept_title=f"Core Concept: {t_name}",
                            explanation=f"This topic explains the mathematical and logical basis of {t_name} as per NCERT syllabus.",
                            key_formulas="Formula 1;Formula 2",
                            misconceptions="Confusing terms;Incorrect order of operations",
                            visual_type="",
                            visual_data="",
                            notes=""
                        )
                    )

                    # 4. Homework Rows
                    resolved_homework.append(
                        HomeworkRow(
                            homework_id=f"HW_CH{ch_no:02d}_T{t_idx:02d}_001",
                            chapter_id=ch_id,
                            topic_id=t_id,
                            set_title=f"Practice: {t_name}",
                            sequence_no=1,
                            question_text=f"Solve the basic practice problem for {t_name}.",
                            marks=2,
                            difficulty="Easy",
                            answer="Calculated solution",
                            explanation="Step-by-step resolution outline.",
                            status="active"
                        )
                    )
                    resolved_homework.append(
                        HomeworkRow(
                            homework_id=f"HW_CH{ch_no:02d}_T{t_idx:02d}_002",
                            chapter_id=ch_id,
                            topic_id=t_id,
                            set_title=f"Practice: {t_name}",
                            sequence_no=2,
                            question_text=f"Solve the application problem for {t_name}.",
                            marks=3,
                            difficulty="Medium",
                            answer="Advanced calculation solution",
                            explanation="Detailed application step analysis.",
                            status="active"
                        )
                    )

                    # 5. Resource Rows
                    resolved_resources.append(
                        ResourceRow(
                            resource_id=f"RES_CH{ch_no:02d}_T{t_idx:02d}_SB",
                            chapter_id=ch_id,
                            topic_id=t_id,
                            resource_type="smart_board",
                            title=f"Smart Board Presentation: {t_name}",
                            url="https://docs.google.com/presentation/d/placeholder",
                            description="Interactive class presentation slides",
                            status="active"
                        )
                    )

                    global_seq += 1

    wb = Workbook()
    wb.remove(wb.active)

    # Write the 7 sheets in order
    _write_custom_sheet(wb, "Chapter_Map", CHAPTER_MAP_HEADERS, resolved_chapters or [])
    _write_custom_sheet(wb, "Topic_Map", TOPIC_MAP_HEADERS, resolved_topics or [])
    _write_custom_sheet(wb, "Lesson_Plans", LESSON_PLANS_HEADERS, resolved_lessons or [])
    _write_custom_sheet(wb, "Concepts", CONCEPTS_HEADERS, resolved_concepts or [])
    _write_custom_sheet(wb, "Homework", HOMEWORK_HEADERS, resolved_homework or [])
    _write_custom_sheet(wb, "Resources", RESOURCES_HEADERS, resolved_resources or [])
    _write_questions_sheet(wb, rows)

    for forbidden in FORBIDDEN_SHEETS:
        if forbidden in wb.sheetnames:
            raise ValueError(f"Refusing to write forbidden sheet: {forbidden}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return errors


def _write_questions_sheet(wb: Workbook, rows: list[QuestionRow]) -> None:
    ws = wb.create_sheet(QUESTIONS_SHEET_NAME)
    ws.append(SAH_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row.to_sheet_row())

    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(SAH_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        width = 14
        if header in ("Question", "Answer / Solution", "Explanation", "Asset Data"):
            width = 48
        elif header in ("Chapter", "Topic", "Learning Outcome", "NCERT Reference"):
            width = 28
        ws.column_dimensions[letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP

    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(len(SAH_HEADERS))}{ws.max_row}"
        table = Table(displayName="QuestionsTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def _write_custom_sheet(wb: Workbook, name: str, headers: list[str], rows: list[Any]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row.to_sheet_row())

    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = 16
        if "title" in header or "name" in header or "outcome" in header or "objectives" in header:
            width = 28
        elif "phase_" in header or "explanation" in header or "text" in header or "url" in header or "data" in header:
            width = 48
        ws.column_dimensions[letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP

    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        table = Table(displayName=f"{name}Table", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
