#!/usr/bin/env python3
"""Validate a subject courseware Excel workbook (7 sheets) against SAH rules."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

_AGENT_DIR = Path(__file__).resolve().parent
if str(_AGENT_DIR) not in sys.path:
    sys.path.insert(0, str(_AGENT_DIR))

from schema import FORBIDDEN_EXTRA_SHEETS, QUESTIONS_SHEET_NAME, SAH_HEADERS

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install dependencies: pip install -r tools/question_bank_agent/requirements.txt", file=sys.stderr)
    sys.exit(2)

# Column schemas
CHAPTER_MAP_HEADERS = ["chapter_id", "chapter_no", "chapter_title", "default_priority", "status"]
TOPIC_MAP_HEADERS = ["topic_id", "chapter_id", "sequence_no", "topic_title", "relative_weight", "relative_difficulty", "learning_outcomes", "status"]
LESSON_PLANS_HEADERS = ["lesson_plan_id", "chapter_id", "topic_id", "objectives", "phase_engage", "phase_explore", "phase_explain", "phase_elaborate", "phase_evaluate", "required_resources", "notes"]
CONCEPTS_HEADERS = ["concept_id", "chapter_id", "topic_id", "concept_title", "explanation", "key_formulas", "misconceptions", "visual_type", "visual_data", "notes"]
HOMEWORK_HEADERS = ["homework_id", "chapter_id", "topic_id", "set_title", "sequence_no", "question_text", "marks", "difficulty", "answer", "explanation", "status"]
RESOURCES_HEADERS = ["resource_id", "chapter_id", "topic_id", "resource_type", "title", "url", "description", "status"]

REQUIRED_SHEETS = [
    "Chapter_Map",
    "Topic_Map",
    "Lesson_Plans",
    "Concepts",
    "Homework",
    "Resources",
    QUESTIONS_SHEET_NAME
]

MATHS_CASE_PARTS = ("(i)", "(ii)", "(iii)")
MATHS_CASE_MARKS = (1, 1, 2)


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_sheet_headers(sheet_name: str, actual_headers: list[str], expected_headers: list[str], errors: list[str]) -> bool:
    if actual_headers != expected_headers:
        if len(actual_headers) != len(expected_headers):
            errors.append(f"Sheet {sheet_name!r}: Header count {len(actual_headers)} != expected {len(expected_headers)}")
        for i, (got, want) in enumerate(zip(actual_headers, expected_headers)):
            if got != want:
                errors.append(f"Sheet {sheet_name!r} Column {i + 1}: got {got!r}, want {want!r}")
                break
        return False
    return True


def validate_workbook(path: Path) -> list[str]:
    errors: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)

    # 1. Check for forbidden sheets
    for name in FORBIDDEN_EXTRA_SHEETS:
        if name in wb.sheetnames:
            errors.append(f"Forbidden sheet present: {name!r}")

    # 2. Check for missing required sheets
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet_name!r}")

    if errors:
        wb.close()
        return errors

    # 3. Validate Chapter_Map
    ws_ch = wb["Chapter_Map"]
    ch_rows = list(ws_ch.iter_rows(values_only=True))
    ch_ids = set()
    if not ch_rows:
        errors.append("Chapter_Map sheet is empty")
    else:
        ch_headers = [cell_str(h) for h in ch_rows[0]]
        validate_sheet_headers("Chapter_Map", ch_headers, CHAPTER_MAP_HEADERS, errors)
        
        col_idx = {h: i for i, h in enumerate(ch_headers) if h}
        for idx, row in enumerate(ch_rows[1:], start=2):
            if not any(cell_str(c) for c in row): continue
            ch_id = cell_str(row[col_idx.get("chapter_id", 0)])
            if not ch_id:
                errors.append(f"Chapter_Map Row {idx}: missing chapter_id")
                continue
            if ch_id in ch_ids:
                errors.append(f"Chapter_Map Row {idx}: duplicate chapter_id {ch_id!r}")
            ch_ids.add(ch_id)

    # 4. Validate Topic_Map
    ws_top = wb["Topic_Map"]
    top_rows = list(ws_top.iter_rows(values_only=True))
    topic_ids = set()
    if not top_rows:
        errors.append("Topic_Map sheet is empty")
    else:
        top_headers = [cell_str(h) for h in top_rows[0]]
        validate_sheet_headers("Topic_Map", top_headers, TOPIC_MAP_HEADERS, errors)
        
        col_idx = {h: i for i, h in enumerate(top_headers) if h}
        for idx, row in enumerate(top_rows[1:], start=2):
            if not any(cell_str(c) for c in row): continue
            t_id = cell_str(row[col_idx.get("topic_id", 0)])
            c_id = cell_str(row[col_idx.get("chapter_id", 1)])
            
            if not t_id:
                errors.append(f"Topic_Map Row {idx}: missing topic_id")
                continue
            if t_id in topic_ids:
                errors.append(f"Topic_Map Row {idx}: duplicate topic_id {t_id!r}")
            topic_ids.add(t_id)

            if c_id not in ch_ids:
                errors.append(f"Topic_Map Row {idx} ({t_id}): chapter_id {c_id!r} not found in Chapter_Map")

    # Helper function to check referential integrity for topic-linked sheets
    def validate_topic_linked_sheet(sheet_name: str, expected_headers: list[str], pk_col: str):
        ws = wb[sheet_name]
        ws_rows = list(ws.iter_rows(values_only=True))
        if not ws_rows:
            errors.append(f"{sheet_name} sheet is empty")
            return
        
        actual_headers = [cell_str(h) for h in ws_rows[0]]
        validate_sheet_headers(sheet_name, actual_headers, expected_headers, errors)
        
        col = {h: i for i, h in enumerate(actual_headers) if h}
        seen_pks = set()
        for idx, row in enumerate(ws_rows[1:], start=2):
            if not any(cell_str(c) for c in row): continue
            pk = cell_str(row[col.get(pk_col, 0)])
            t_id = cell_str(row[col.get("topic_id", 2)])
            c_id = cell_str(row[col.get("chapter_id", 1)])

            if not pk:
                errors.append(f"{sheet_name} Row {idx}: missing {pk_col}")
                continue
            if pk in seen_pks:
                # Homework can have multiple questions in same set, so homework_id is primary key
                errors.append(f"{sheet_name} Row {idx}: duplicate {pk_col} {pk!r}")
            seen_pks.add(pk)

            if t_id not in topic_ids:
                errors.append(f"{sheet_name} Row {idx} ({pk}): topic_id {t_id!r} not found in Topic_Map")
            if c_id not in ch_ids:
                errors.append(f"{sheet_name} Row {idx} ({pk}): chapter_id {c_id!r} not found in Chapter_Map")

    # 5. Validate Lesson_Plans
    validate_topic_linked_sheet("Lesson_Plans", LESSON_PLANS_HEADERS, "lesson_plan_id")

    # 6. Validate Concepts
    validate_topic_linked_sheet("Concepts", CONCEPTS_HEADERS, "concept_id")

    # 7. Validate Homework
    validate_topic_linked_sheet("Homework", HOMEWORK_HEADERS, "homework_id")

    # 8. Validate Resources
    validate_topic_linked_sheet("Resources", RESOURCES_HEADERS, "resource_id")

    # 9. Validate Questions (existing validations)
    ws_q = wb[QUESTIONS_SHEET_NAME]
    q_rows = list(ws_q.iter_rows(values_only=True))
    if not q_rows:
        errors.append("Questions sheet is empty")
    else:
        q_headers = [cell_str(h) for h in q_rows[0]]
        validate_sheet_headers(QUESTIONS_SHEET_NAME, q_headers, SAH_HEADERS, errors)
        
        col = {h: i for i, h in enumerate(q_headers) if h}
        seen_qids = set()
        prev_chapter_no = None

        for row_idx, row in enumerate(q_rows[1:], start=2):
            if not any(cell_str(c) for c in row): continue

            def get(header: str) -> str:
                idx = col.get(header)
                if idx is None: return ""
                return cell_str(row[idx]) if idx < len(row) else ""

            qid = get("Question ID")
            if not qid:
                errors.append(f"Questions Row {row_idx}: missing Question ID")
                continue
            if qid in seen_qids:
                errors.append(f"Questions Row {row_idx}: duplicate Question ID {qid!r}")
            seen_qids.add(qid)

            use = get("Use in Papers")
            if use and use != "Yes":
                errors.append(f"Questions Row {row_idx} ({qid}): Use in Papers must be Yes, got {use!r}")

            ch_raw = get("Chapter No.")
            try:
                ch_num = float(ch_raw) if ch_raw else float("nan")
            except ValueError:
                errors.append(f"Questions Row {row_idx} ({qid}): invalid Chapter No. {ch_raw!r}")
                ch_num = float("nan")

            if ch_raw:
                if prev_chapter_no is not None and ch_num < prev_chapter_no:
                    errors.append(
                        f"Questions Row {row_idx} ({qid}): rows not sorted by Chapter No. ({ch_num} after {prev_chapter_no})"
                    )
                prev_chapter_no = ch_num

            qtype = get("Question Type")
            subject = get("Subject")
            marks_raw = get("Marks")
            question_text = get("Question")

            if qtype == "MCQ":
                for opt in ("Option A", "Option B", "Option C", "Option D"):
                    if not get(opt):
                        errors.append(f"Questions Row {row_idx} ({qid}): MCQ missing {opt}")
                if not get("Correct Answer"):
                    errors.append(f"Questions Row {row_idx} ({qid}): MCQ missing Correct Answer")

            if subject == "Maths" and qtype == "Case/Source-Based":
                for part in MATHS_CASE_PARTS:
                    if part not in question_text:
                        errors.append(f"Questions Row {row_idx} ({qid}): Maths case missing sub-part {part}")
                try:
                    marks = int(float(marks_raw)) if marks_raw else 0
                except ValueError:
                    marks = -1
                if marks != sum(MATHS_CASE_MARKS):
                    errors.append(
                        f"Questions Row {row_idx} ({qid}): Maths case marks must be 4 (1+1+2), got {marks_raw!r}"
                    )

            asset_fmt = get("Asset Format")
            asset_data = get("Asset Data")
            image_url = get("Image URL")
            if asset_fmt and not asset_data and not image_url:
                errors.append(f"Questions Row {row_idx} ({qid}): Asset Format set but Asset Data and Image URL empty")

    wb.close()
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate SAH subject courseware Excel workbook")
    parser.add_argument("workbook", type=Path, help="Path to .xlsx file")
    args = parser.parse_args()

    if not args.workbook.is_file():
        print(f"File not found: {args.workbook}", file=sys.stderr)
        return 1

    errors = validate_workbook(args.workbook)
    if errors:
        print(f"FAIL: {args.workbook}")
        for err in errors:
            print(f"  - {err}")
        return 1

    print(f"OK: {args.workbook}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
